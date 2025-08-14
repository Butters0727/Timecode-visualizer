import os
import openpyxl
from datetime import datetime

def get_start_end_time_and_duration(tc_path):
    start_time_str = None
    end_time_str = None
    duration_str = None
    with open(tc_path, 'r') as f:
        lines = f.readlines()
        if lines:
            first_line_values = lines[0].strip().split(',')
            if len(first_line_values) >= 4:
                hour, minute, sec = first_line_values[1:4]
                start_time_str = f"{hour}:{minute}:{sec}"

            last_line_values = lines[-1].strip().split(',')
            if len(last_line_values) >= 4:
                hour, minute, sec = last_line_values[1:4]
                end_time_str = f"{hour}:{minute}:{sec}"
            
            if start_time_str and end_time_str:
                t1 = datetime.strptime(start_time_str, "%H:%M:%S")
                t2 = datetime.strptime(end_time_str, "%H:%M:%S")
                duration = t2 - t1
                duration_str = str(duration)

    return start_time_str, end_time_str, duration_str

def get_frame_count(bvh_path):
    with open(bvh_path, 'r') as f:
        for line in f:
            if line.startswith("Frames:"):
                return int(line.split(":")[1].strip())
    return None

def main():
    bvh_folder = 'BVH'
    tc_folder = 'TC'
    output_file = 'data.xlsx'

    try:
        wb = openpyxl.load_workbook(output_file)
        ws = wb.active
    except FileNotFoundError:
        print(f"Error: '{output_file}' not found.")
        return

    header = [cell.value for cell in ws[1]]
    try:
        file_col = header.index("BVH")
        start_time_col = header.index("TIME START")
        end_time_col = header.index("TIME END")
        duration_col = header.index("DURATION")
        frames_col = header.index("FRAMES")
    except ValueError as e:
        print(f"Error: Missing column in excel file - {e}")
        return

    for row in ws.iter_rows(min_row=2):
        file_name = row[file_col].value
        if file_name:
            base_name, _ = os.path.splitext(file_name)
            bvh_path = os.path.join(bvh_folder, base_name + ".bvh")
            tc_path = os.path.join(tc_folder, base_name + ".tc")

            if os.path.exists(bvh_path) and os.path.exists(tc_path):
                frames = get_frame_count(bvh_path)
                start_time, end_time, duration = get_start_end_time_and_duration(tc_path)

                if frames is not None:
                    row[frames_col].value = frames
                if start_time is not None:
                    row[start_time_col].value = start_time
                if end_time is not None:
                    row[end_time_col].value = end_time
                if duration is not None:
                    row[duration_col].value = duration

    wb.save(output_file)
    print(f"Data written to {output_file}")

if __name__ == "__main__":
    main()